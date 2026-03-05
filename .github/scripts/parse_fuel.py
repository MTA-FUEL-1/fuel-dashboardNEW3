import os, json, glob
from openpyxl import load_workbook
from collections import defaultdict

# Look everywhere for xlsx files
xlsx_files = sorted(
    glob.glob('data/xlsx/*.xlsx') +
    glob.glob('data/*.xlsx') +
    glob.glob('*.xlsx')
)

print(f"Found {len(xlsx_files)} xlsx files: {xlsx_files}")

for xlsx_path in xlsx_files:
    basename = os.path.basename(xlsx_path)
    # Extract date from filename - look for MM.DD.YYYY pattern
    import re
    date_match = re.search(r'(\d{2}\.\d{2}\.\d{4})', basename)
    if not date_match:
        print(f"Could not extract date from {basename}")
        continue
    
    date_str = date_match.group(1)
    json_path = f'data/{date_str}_fuel.json'

    if os.path.exists(json_path):
        try:
            existing = json.load(open(json_path))
            if 'ss' in existing and len(existing['ss']) > 0:
                print(f"Skipping {json_path} - already has full data")
                continue
        except:
            pass

    print(f"Processing {xlsx_path} -> {json_path}")

    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    header_row = None
    for i, row in enumerate(rows):
        row_vals = [str(c).strip().lower() if c else '' for c in row]
        if any('state' in v for v in row_vals) and any('retail' in v for v in row_vals):
            header_row = i
            break

    if header_row is None:
        print(f"Could not find header in {xlsx_path}")
        continue

    headers = [str(c).strip().lower() if c else '' for c in rows[header_row]]

    def find_col(keywords):
        for i, h in enumerate(headers):
            if any(k in h for k in keywords):
                return i
        return None

    col_state  = find_col(['state'])
    col_city   = find_col(['city'])
    col_store  = find_col(['store'])
    col_retail = find_col(['retail'])
    col_disc   = find_col(['discount','disc'])

    if None in [col_state, col_retail, col_disc]:
        print(f"Missing columns in {xlsx_path}: state={col_state} retail={col_retail} disc={col_disc}")
        continue

    state_data = defaultdict(list)

    for row in rows[header_row + 1:]:
        if not row or not row[col_state]:
            continue
        state = str(row[col_state]).strip().upper()
        if len(state) != 2:
            continue
        try:
            retail = float(row[col_retail]) if row[col_retail] else None
            disc   = float(row[col_disc])   if row[col_disc]   else None
            save   = round(retail - disc, 4) if retail and disc else 0
            city   = str(row[col_city]).strip() if col_city and row[col_city] else ''
            store  = str(row[col_store]).strip() if col_store and row[col_store] else ''
            if retail and disc:
                state_data[state].append({
                    'store_no': store,
                    'city': city,
                    'retail': retail,
                    'disc': disc,
                    'save': save
                })
        except:
            continue

    ss = []
    stores = {}
    for state, locs in sorted(state_data.items()):
        retails = [l['retail'] for l in locs]
        discs   = [l['disc']   for l in locs]
        saves   = [l['save']   for l in locs]
        ss.append({
            'state':      state,
            'avg_retail': round(sum(retails)/len(retails), 4),
            'avg_disc':   round(sum(discs)/len(discs), 4),
            'avg_save':   round(sum(saves)/len(saves), 4),
            'best':       round(min(discs), 4),
            'worst':      round(max(discs), 4),
            'n':          len(locs)
        })
        stores[state] = locs

    all_retails = [l['retail'] for locs in state_data.values() for l in locs]
    all_discs   = [l['disc']   for locs in state_data.values() for l in locs]

    nat = {
        'retail': round(sum(all_retails)/len(all_retails), 4),
        'disc':   round(sum(all_discs)/len(all_discs), 4),
        'save':   round((sum(all_retails)-sum(all_discs))/len(all_retails), 4)
    }

    parts = date_str.split('.')
    short_date = f"{parts[0]}/{parts[1]}"

    # Make sure data folder exists
    os.makedirs('data', exist_ok=True)

    output = {
        'date':      date_str,
        'shortDate': short_date,
        'nat':       nat,
        'ss':        ss,
        'stores':    stores
    }

    with open(json_path, 'w') as f:
        json.dump(output, f, indent=2)

    print(f"Generated {json_path} with {len(ss)} states")
